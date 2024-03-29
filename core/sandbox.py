import code
import io
import sys
from pathlib import Path

from utils.common import SandboxResponse
from utils.enumeration import *


class Sandbox:
    def __init__(self) -> None:
        self.interpreter = code.InteractiveInterpreter()
        self.code_history = []
        self.stdout = []
        self.stderr = []
        self.import_lib()

    def import_lib(self):
        code_import = (
            "import openpyxl\nimport pandas as pd\nimport matplotlib.pyplot as plt\nimport os\nimport datetime\n"
        )

        self.step(code_import, dummy=False)

    def load_workbook(self, workbook_path):

        code_init = [f'wb_path = r"{workbook_path}"']
        code_init += [f"workbook = openpyxl.load_workbook(wb_path)"]

        self.step("\n".join(code_init), dummy=False)

    def load_worksheets(self, sheet_vars):
        sheet_names = self.get_existing_sheet_names()

        code_init = []
        for sheet_var, sheet_name in zip(sheet_vars, sheet_names):
            code_init.append(f'{sheet_var} = workbook["{sheet_name}"]')

        self.step("\n".join(code_init), dummy=False)

    def get_existing_sheet_names(self):
        code_snippet = "print(workbook.sheetnames)"
        sheet_names = self.step(code_snippet, dummy=True).msg.splitlines()[-1]
        sheet_names = eval(sheet_names)
        return sheet_names

    def get_sheet_state(self) -> str:
        sheet_state = ""
        sheet_names = self.get_existing_sheet_names()
        for sheet_name in sheet_names:
            code_snippet = f"""print(workbook["{sheet_name}"].max_column)
print(workbook["{sheet_name}"].max_row)
print(workbook["{sheet_name}"].min_column)
print(workbook["{sheet_name}"].min_row)
print(workbook["{sheet_name}"].cell(1, 1).value)"""
            max_column, max_row, min_column, min_row, first_value = self.step(code_snippet, dummy=True).msg.splitlines()
            if (max_column == max_row == min_column == min_row == str(1)) and str(first_value) == "None":
                sheet_desc = 'Sheet "{sheet_name}" is empty. '.format(sheet_name=sheet_name)
            else:
                sheet_desc = 'Sheet "{sheet_name}" has {n_rows} rows (Including the header row) and {n_cols} columns ({headers}). '
                code_snippet = f"""print(workbook["{sheet_name}"].max_column)
print(workbook["{sheet_name}"].max_row)
print([cell.value for cell in workbook["{sheet_name}"][1]])
print([str(cell.value.__class__) for cell in workbook["{sheet_name}"][2]])"""

                n_cols, n_rows, headers, data_types = self.step(code_snippet, dummy=True).msg.splitlines()
                import datetime

                headers = eval(headers)
                data_types = eval(data_types)

                headers_str = ", ".join(
                    [
                        f'{chr(65 + i)}({i+1}): "{header}" ({data_type})'
                        for i, (header, data_type) in enumerate(zip(headers, data_types))
                    ]
                )
                sheet_desc = sheet_desc.format(
                    sheet_name=sheet_name,
                    n_cols=int(n_cols),
                    headers=headers_str,
                    n_rows=int(n_rows),
                )

            sheet_state += sheet_desc

        return sheet_state

    def reset(self):
        self.interpreter = code.InteractiveInterpreter()

    def step(self, code_snippet: str, dummy=False) -> SandboxResponse:
        out_buffer = io.StringIO()
        err_buffer = io.StringIO()
        sys.stdout = out_buffer
        sys.stderr = err_buffer

        self.interpreter.runcode(code_snippet)

        output = out_buffer.getvalue()
        error = err_buffer.getvalue()

        if not dummy:
            if error == "":
                self.code_history.append(code_snippet)
            self.stdout.append(output)
            self.stderr.append(error)

        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__

        if error != "":  # error caught
            # to clear error context
            self.reset()
            self.step("\n".join(self.code_history), dummy=True)
            return SandboxResponse(EXEC_CODE.FAIL, error)

        return SandboxResponse(EXEC_CODE.SUCCESS, output)

    def save(self, save_dir: Path):
        self.step(f'workbook.save(r"{save_dir / "workbook_new.xlsx"}")', dummy=False)
        with open(save_dir / "code.py", "w", encoding="utf-8") as f:
            f.write("\n\n# ============\n".join(self.code_history))

        with open(save_dir / "outputs.txt", "w", encoding="utf-8") as f:
            f.write("\n\n# ============\n".join(self.stdout))

        with open(save_dir / "errors.txt", "w", encoding="utf-8") as f:
            f.write("\n\n# ============\n".join(self.stderr))

    def save_temp_workbook(self, save_dir: Path):
        self.step(f'workbook.close()\nworkbook.save(r"{save_dir / "workbook_temp.xlsx"}")', dummy=True)
